import pandas as pd
from pathlib import Path
from datetime import datetime
from utils import log, get_connection, safe_int, safe_datetime, safe_str, safe_bool, safe_float
import os
import psycopg2
from openpyxl import load_workbook
import shutil
from extract_drawings import getDrawing



def print_job_orders(target_batch_id=None):
    
    job_order_query = """ 
WITH details AS (
SELECT p2.id, p2.ma_number, p.parent_id, p.dimension_x, p.dimension_y, p.dimension_z, p.color_id, p.material_id, c.color_name_bg, m.material_code, p2."name", p2.type_id, p2.description, pt.type_code, m.material_name_bg, p.process_cut, p.process_mill, p2.is_assembly from parts p
LEFT JOIN products p2 ON p2.id = p.product_id
LEFT JOIN colors c ON c.id = p.color_id
LEFT JOIN materials m ON m.id = p.material_id
LEFT JOIN product_type pt ON pt.id = p2.type_id
WHERE p.parent_id is NULL
ORDER BY p2.ma_number ASC
)


SELECT jo.job_order_number, jo.batch_id, jo.quantity, details.*, jo.price_offer_num, pgp.dimension_x_g, pgp.dimension_y_g, pgp.dimension_z_g, pgp.qty_per_group, pgp.id, pj.project_name from job_orders jo
LEFT JOIN details ON details.id = jo.product_id
left join parts_group_process pgp on pgp.product_id = jo.product_id
left join product_projects pp on pp.product_id  = jo.product_id
left join projects pj on pj.id = pp.project_id 
WHERE jo.batch_id = %s
ORDER BY jo.job_order_number ASC"""

    query = """
    SELECT 
        p2.id,
        p2.ma_number,
        p2.name,
        p2.description,
        p2.is_assembly,
        p.id as part_id,
        p.part_number,
        p.dimension_x,
        p.dimension_y,
        p.dimension_z,
        p.process_cut,
        p.process_mill,
        c.color_name_bg,
        m.material_name_bg,
        bom.quantity,
        p.parent_id
    FROM parts p
    JOIN bill_of_materials bom ON bom.part_id = p.id
    JOIN products p2 ON p2.id = bom.product_id
    LEFT JOIN colors c ON c.id = p.color_id
    LEFT JOIN materials m ON m.id = p.material_id
    WHERE p2.is_assembly = TRUE 
        AND p.parent_id = %s
    ORDER BY p2.ma_number ASC
"""


    
    if not target_batch_id:
        log("Batch ID is required to print job orders", "ERROR")
        return False
    
    log(f"📋 Printing job orders for batch_id: {target_batch_id}")
    conn = get_connection()
    if not conn:
        return False
    
    try:
        # Create batch directory using batch_id
        BASE_OUTPUT_PATH = Path("S:/Exchange_Folder/JO/Batches")
        BASE_OUTPUT_PATH_JO = Path("S:/JO")
        
        batch_dir = BASE_OUTPUT_PATH / str(target_batch_id)
        batch_dir.mkdir(parents=True, exist_ok=True)
        log(f"Working directory: {batch_dir.absolute()}")
        
        # Path to template file
        template_path = Path("jo.xlsx")
        appended_template_path = Path("appendix.xlsx")  # "appendix.xlsx"
        
        if not template_path.exists():
            log(f"Template file not found: {template_path}", "ERROR")
            return False
        if not appended_template_path.exists():
            log(f"Appended template file not found: {appended_template_path}", "ERROR")
            return False
        
        
        cursor = conn.cursor()
        cursor.execute(job_order_query, (target_batch_id,))
        # cursor.execute(job_order_query.format(target_batch_id=target_batch_id))//SQL injection risk if target_batch_id is not sanitized
        rows = cursor.fetchall()
        
        
        
        if not rows:
            log("No data returned from query", "WARNING")
            return False
        
        success_count = 0
        for idx, row in enumerate(rows, 1):
            try:
                job_order_number = safe_str(row[0]) if row[0] else f"JO_{idx}"
                batch_id = safe_str(row[1]) if row[1] else target_batch_id
                quantity = safe_int(row[2]) if row[2] else 0
                if batch_id != str(target_batch_id):
                    log(f"   ⚠️ Missing batch_id: {batch_id}", "WARNING")
                    break
                id = safe_str(row[3]) if row[3] else ""
                ma_number = safe_str(row[4]) if row[4] else ""
                dimension_x = safe_float(row[6]) if row[6] else ""
                dimension_y = safe_float(row[7]) if row[7] else ""
                dimension_z = safe_float(row[8]) if row[8] else ""  
                color_name_bg = safe_str(row[11]) if row[11] else ""
                name = safe_str(row[13]) if row[13] else ""
                description = safe_str(row[15]) if row[15] else ""
                type_code = safe_str(row[16]) if row[16] else ""
                material_name_bg = safe_str(row[17]) if row[17] else ""
                process_cut = safe_bool(row[18]) if row[18] else False
                process_mill = safe_bool(row[19]) if row[19] else False
                is_assembly = safe_bool(row[20]) if row[20] else False
                price_offer_num = safe_str(row[21]) if row[21] else ""
                dimension_x_g = safe_float(row[22]) if row[22] else ""
                dimension_y_g = safe_float(row[23]) if row[23] else ""
                dimension_z_g = safe_float(row[24]) if row[24] else ""
                qty_per_group = safe_int(row[25]) if row[25] else 0
                group_id = safe_int(row[26]) if row[26] else 0
                project_name = safe_str(row[27]) if row[27] else ""
                
                
                
                search_string = ma_number[6:]
         
                matyching_files = getDrawing(root_path=r'H:\drawings', keyword=search_string, extension=".pdf")
                print(f"   🔍 Found {len(matyching_files)} matching drawing files for MA_Number {search_string}")  
                # After finding matching files
                if matyching_files:
                    # Create drawings subdirectory
                    drawings_dir = batch_dir / "drawings"
                    drawings_dir.mkdir(exist_ok=True)
                    
                    for drawing_file in matyching_files:
                        src_path = Path(drawing_file)
                        dst_path = drawings_dir / src_path.name
                        shutil.copy2(src_path, dst_path)
                    
                    log(f"   📄 Copied {len(matyching_files)} drawing(s) to {drawings_dir}")
                
                 
                # Use job order number as filename
                filename = f"{job_order_number}.xlsx"
                output_path = batch_dir / filename
                
                # Copy template to new file
                shutil.copy(template_path, output_path)
                
                # Open the workbook
                wb = load_workbook(output_path)
                sheet = wb.active
                
                # Just put data in cells - adjust cell references as needed
                sheet['J16'] = job_order_number
                sheet['E11'] = batch_id
                sheet['B38'] = id
                sheet['D38'] = ma_number
                sheet['H38'] = color_name_bg
                sheet['E38'] = material_name_bg
                sheet['D6'] = datetime.now().strftime("%Y-%m-%d")
                sheet['D16'] = datetime.now().strftime("%Y-%m-%d")
                sheet['D22'] = quantity
                sheet['B18'] = project_name
                
                if group_id and qty_per_group:
                    sheet['N27'] = dimension_x_g
                    sheet['N28'] = dimension_y_g
                    sheet['N29'] = dimension_z_g
                    sheet['J20'] = round(quantity % qty_per_group) if quantity % qty_per_group else quantity /qty_per_group
                    
                else:   
                    sheet['N27'] = dimension_x
                    sheet['N28'] = dimension_y
                    sheet['N29'] = dimension_z
                    sheet['J20'] = quantity
                    
                    
                
                
                
                sheet['I38'] = type_code
                # sheet['L11'] = description
                sheet['D10'] = name
                sheet['G11'] = price_offer_num
                
                
                
                if process_cut and process_mill:
                    sheet['L21'] = "                   "
                elif process_cut:
                    sheet['L21'] = " ⚠️ РЯЗАНЕ         "
                elif process_mill:
                    sheet['L21'] = " ⚠️ ФРЕЗОВАНЕ       "
                    
                    
                if id and is_assembly == True:
                    print(f"   🔍 Fetching child parts for product_id {id}")
                    cursor.execute(query, (id,))
                    row_childrens = cursor.fetchall()
                    print(len(row_childrens))
                    sheet['S21'] = len(row_childrens)
                    
    
                    if row_childrens:
                        child_info_set = add_child_parts(row_childrens, id)
        
                        # Create appendix file for child parts
                        filename_appendix = f"{job_order_number}_append_{id}.xlsx"
                        output_path_appendix = batch_dir / filename_appendix
                        shutil.copy(appended_template_path, output_path_appendix)
                        wb_appendix = load_workbook(output_path_appendix)
                        sheet_appendix = wb_appendix.active
                        
                        sheet_appendix['I3'] = f"{job_order_number}"
                        
        
                        for idx_child, child_info in enumerate(child_info_set.values(), 1):
                            row_num = idx_child + 5  # Start from row 6 in appendix
                            sheet_appendix[f"B{row_num}"] = child_info.get("part_id", "")
                            sheet_appendix[f"C{row_num}"] = child_info.get("ma_number", "")
                            sheet_appendix[f"D{row_num}"] = child_info.get("child_ma_number", "")
                            sheet_appendix[f"E{row_num}"] = child_info.get("child_dimension_x", "")
                            sheet_appendix[f"F{row_num}"] = child_info.get("child_dimension_y", "")
                            sheet_appendix[f"G{row_num}"] = child_info.get("child_dimension_z", "")
                            sheet_appendix[f"I{row_num}"] = child_info.get("child_color_name_bg", "")
                            sheet_appendix[f"J{row_num}"] = child_info.get("child_material_name_bg", "")
                            child_quantity = int(child_info.get("child_quantity", 0)) if child_info.get("child_quantity", 0) else 0
                            sheet_appendix[f"K{row_num}"] = child_quantity * quantity  # Total quantity needed based on parent quantity
                            sheet_appendix[f"H{row_num}"] = child_info.get("process_cut", "") or child_info.get("process_mill", "")
                            
                        wb_appendix.save(output_path_appendix)
                        wb_appendix.close()
                        log(f"   ✅ Created appendix: {output_path_appendix}")
                    else:
                        log(f"   ⚠️ No child parts found for assembly {id}", "WARNING")

                    # Save the main workbook (ALWAYS)
                wb.save(output_path)
                wb.save(BASE_OUTPUT_PATH_JO / filename)  # Save a copy in the JO folder
                success_count += 1
                log(f"✅ [{idx}/{len(rows)}] Created main: {output_path}")
                wb.close()
                
            except Exception as e:
                log(f"❌ Error processing job order {row[0] if row else 'unknown'}: {str(e)}", "ERROR")
                continue
        
        cursor.close()
        conn.close()
        
        log(f"✅ Successfully created {success_count} out of {len(rows)} job order files in {batch_dir}")
        return success_count > 0
        
    except Exception as e:
        log(f"❌ Error in print_job_orders: {str(e)}", "ERROR")
        if conn:
            conn.close()
        return False

def add_child_parts(row_childrens, id):
    print(f"   🔍 Found {len(row_childrens)} child parts for product_id {id}")
    
    child_info_set = {}
    for row_child in row_childrens:
        # Use proper indexing based on the query order
        # Index mapping:
        # 0: p2.id
        # 1: p2.ma_number
        # 2: p2.name
        # 3: p2.description
        # 4: p2.is_assembly
        # 5: part_id
        # 6: part_number
        # 7: dimension_x
        # 8: dimension_y
        # 9: dimension_z
        # 10: process_cut
        # 11: process_mill
        # 12: color_name_bg
        # 13: material_name_bg
        # 14: quantity
        # 15: parent_id
        
        product_id = safe_int(row_child[0]) if row_child[0] else None
        ma_number = safe_str(row_child[1]) if row_child[1] else "N/A"
        product_name = safe_str(row_child[2]) if row_child[2] else "N/A"
        part_id = safe_int(row_child[5]) if row_child[5] else None
        part_number = safe_str(row_child[6]) if row_child[6] else "N/A"
        child_dimension_x = safe_float(row_child[7]) if row_child[7] else 0.0
        child_dimension_y = safe_float(row_child[8]) if row_child[8] else 0.0
        child_dimension_z = safe_float(row_child[9]) if row_child[9] else 0.0
        process_cut = safe_bool(row_child[10]) if len(row_child) > 10 else False
        process_mill = safe_bool(row_child[11]) if len(row_child) > 11 else False
        child_color_name_bg = safe_str(row_child[12]) if len(row_child) > 12 and row_child[12] else "N/A"
        child_material_name_bg = safe_str(row_child[13]) if len(row_child) > 13 and row_child[13] else "N/A"
        child_quantity = safe_int(row_child[14]) if len(row_child) > 14 and row_child[14] else 0
        parent_id = safe_int(row_child[15]) if len(row_child) > 15 and row_child[15] else None
        
        # Use part_id as key, or part_number if part_id is None
        key = str(part_id) if part_id else part_number
        
        child_info_set[key] = {
            "product_id": product_id,
            "ma_number": ma_number,
            "product_name": product_name,
            "part_id": part_id,
            "child_ma_number": part_number,
            "child_dimension_x": child_dimension_x,
            "child_dimension_y": child_dimension_y,
            "child_dimension_z": child_dimension_z,
            "child_color_name_bg": child_color_name_bg,
            "child_material_name_bg": child_material_name_bg,
            "child_quantity": child_quantity,
            "parent_id": parent_id
        }
        
        if process_cut:
            print(f"   🔍 Child part {part_number} requires cutting process")
            child_info_set[key]["process_cut"] = "РЯЗАНЕ"
        if process_mill:
            print(f"   🔍 Child part {part_number} requires milling process")
            child_info_set[key]["process_mill"] = "ФРЕЗОВАНЕ"
    
    return child_info_set



        
        